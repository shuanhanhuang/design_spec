# from django.shortcuts import render

# Create your views here.
import json
from django.conf import settings
from django.shortcuts import render, redirect
from specapp.form import SpecForm,CompareForm,productForm,tablespecForm
from specapp.models import Spec,Compare,product,Special,tablespec
from django.http import HttpResponse
# from specapp.form import SpecForm,CompareForm,OrderBOMForm,productForm,tableBOMForm,tablespecForm
# from specapp.models import Spec,Compare,OrderBOM,product,tableBOM,Special,tablespec
from django.db.models import Q
import pdfplumber
import re
count = 0
import requests
import openpyxl
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from django.template.loader import get_template
from django.template.loader import render_to_string
from excel_response import ExcelResponse
from xhtml2pdf import pisa
import io
def font_patch():
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    from xhtml2pdf.default import DEFAULT_FONT
    pdfmetrics.registerFont(TTFont('yh', 'D:/冷氣廠規格書/規範書系統/design_spec/static/fonts/msyh.ttf'))
    DEFAULT_FONT['helvetica'] = 'yh'

def download_workbook(request,cNumber=None):
    CompareSpec = Spec.objects.get(cNumber=cNumber)
    excel_filter = Compare.objects.filter(ref=CompareSpec)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="compare.xlsx"'

    wb = Workbook()
    ws = wb.active
    columns = ['筆數','契約規範', '送審規格', '是否符合', '頁碼', '廠牌','備註']
    ws.append(columns)
    data = excel_filter
    print(data[0].ctable1)
    for index,row in enumerate(data):
        row_data = [index+1,row.ctable1, row.ctable2, row.cConform, row.cPage,row.cBrand,row.cRemark]  # Adjust fields as needed
        ws.append(row_data)

    wb.save(response)
    return response


def render_pdf_view(request,cNumber=None):
    spec = Spec.objects.get(cNumber=cNumber)
    Tablespec = tablespec.objects.get(cNumber=cNumber)
    context = {'Tablespec':Tablespec,'spec':spec}  # 這裡填入您的模板上下文
    html_content = render_to_string('pdfconvert.html', context)
    font_patch()
    
    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_content.encode('utf-8'), dest=pdf_file, encoding='utf-8')
    
    if not pisa_status.err:
        pdf_file.seek(0)
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="export.pdf"'
        return response
    
    return HttpResponse('Error converting HTML to PDF', status=500)

def designPost(request):
    current_time = str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    # if 'user_id' in request.session:  # 假设你的用户 ID 存在于 session 中
    #     user_id = request.session['user_id']
    #     user = User.objects.get(pk=user_id)  # 获取当前登录用户信息
    #     firstname = user.name  # 传递用户名到模板中
    #     position = user.Position
    #     title = user.title
    if request.method == "POST":  #如果是以POST方式才處理
        homeform = SpecForm(request.POST,request.FILES)  #建立forms物件
        if homeform.is_valid():
            cNumber= homeform.cleaned_data['cNumber']
            cProject = homeform.cleaned_data['cProject'] #取得表單輸入資料
            cName = homeform.cleaned_data['cName']
            cClient = homeform.cleaned_data['cClient']
            clocation = homeform.cleaned_data['clocation']
            cContact = homeform.cleaned_data['cContact']
            cContactTel = homeform.cleaned_data['cContactTel']
            cDate = current_time
            cDeliveryDate = homeform.cleaned_data['cDeliveryDate']
            cFile = homeform.cleaned_data['cFile']
            chaveorder = "無"
            unit = Spec.objects.create(cNumber=cNumber,cProject=cProject, cName=cName, cClient=cClient,clocation=clocation,cContact=cContact,cContactTel=cContactTel,cDate=cDate,cDeliveryDate=cDeliveryDate,cFile=cFile,chaveorder=chaveorder)
            home = unit.save()  #寫入資料庫
            message = '已儲存...'
            return redirect('/designIndex/')

        else:
            message = "每個欄位都必須填寫"
    else:
        message = '編號、勝新業務、案件名稱、客戶名稱、工地名稱必須填寫!!!'
    homeform = SpecForm()
    return render(request, "designPost.html", locals())

def designEdit(request,id=None,mode=None):
    homeform = SpecForm(request.POST)
    unitspec = Spec.objects.get(id = id)
    if mode == "load":  # 由 index.html 按 編輯二 鈕
        return render(request, "designEdit.html", locals())
    elif mode == "save": # 由 edit2.html 按 submit
        unitspec = Spec.objects.get(id = id)
        unitspec.clocation = request.POST['clocation']
        unitspec.cContact = request.POST['cContact']
        unitspec.cContactTel = request.POST['cContactTel']
        unitspec.cDeliveryDate = request.POST['cDeliveryDate']
        unitspec.save()
        message = '已修改...'
        return redirect('/designIndex/')
    
def designView(request,id=None):
    home = Spec.objects.get(id=id)
    return render(request, "designView.html",locals())

def designIndex(request):
    all = Spec.objects.all().order_by("id")
    allHomeCount = len(all)
    return render(request, "designIndex.html",locals())

def designDelete(request,id=None):
	if id!=None:
		if request.method == "POST":  #如果是以POST方式才處理
			id=request.POST['cId'] #取得表單輸入的編號
		try:
			unit = Spec.objects.get(id=id)
			unit.delete()
			return redirect('/designIndex/')
		except:
			message = "讀取錯誤!"
	return render(request, "designDelete.html", locals())

def fileUpload(request,id=None,mode=None):
    homeform = SpecForm(request.POST)
    if mode == "load":  # 由 index.html 按 編輯二 鈕
        unit = Spec.objects.get(id=id)  #取得要修改的資料記  
        return render(request, "fileUpload.html", locals())
    elif mode == "save": # 由 edit2.html 按 submit
        unit = Spec.objects.get(id=id)  #取得要修改的資料記錄
        try:
            unit.cFile = request.FILES['cFile']
        except:
            unit.cFile = request.POST["cFile"]
        unit.save()  #寫入資料庫
        message = '已修改...'
        return redirect('/designIndex/')
    

    
def comparePost(request,table1,table2,cNumber):
    CompareSpec = Spec.objects.get(cNumber=cNumber)
    unitinner = Compare.objects.filter(ref=CompareSpec).order_by("id")
    Compareform = CompareForm(request.POST)  #建立forms物件
    if Compareform.is_valid():
        cNumber = CompareSpec.cNumber
        ctable1= table1
        ctable2 = table2 #取得表單輸入資料
        cConform = "符合"
        cPage = "無"
        cBrand = "勝新"
        cRemark = Compareform.cleaned_data['cRemark']  
        unit = Compare.objects.create(ref=CompareSpec,cNumber=cNumber,ctable1=ctable1,ctable2=ctable2, cConform=cConform, cPage=cPage, cBrand=cBrand, cRemark=cRemark)
        home = unit.save()  #寫入資料庫
        message = '已儲存...'
        return True


def compareIndex(request,cNumber=None):
    n = cNumber
    CompareSpec = Spec.objects.get(cNumber=cNumber)
    aaa = Compare.objects.filter(ref=CompareSpec)
    try:
        c = aaa[0].ctable1
    except:
        pdf = pdfplumber.open("./media/"+str(CompareSpec.cFile))
        text = ""
        for page in pdf.pages:
            page_text = page.extract_text()
            page_text = re.sub(r'\d+ \|', "", page_text)
            page_text = re.sub(r'第 \d+ 章\n空氣調節箱|第 \d+ 章\n空氣調節箱', "", page_text)
            text += page_text
        sections = re.split(r'\n(?=\d+\.\d+\s+|\d+\.\s+|\(\d+\)\s+|2.\d.\d\s+)', text)
        a = [section.strip() for section in sections if section.strip()]
        for table1 in a:
            table1 = table1.replace(' ',"")
            table2 = table1
            table2 = table2.replace('必需',"")
            table2 = table2.replace('必須',"")
            table2 = table2.replace('需',"")
            table2 = table2.replace('須',"")
            table2 = table2.replace('應',"")
            table2 = table2.replace('至少',"")
            table2 = table2.replace('並',"")
            table2 = table2.replace('確保',"")
            choose = comparePost(request,table1,table2,cNumber)
    all = Compare.objects.filter(ref=CompareSpec).order_by("id")
    allHomeCount = len(all)
    return render(request, "compareIndex.html",locals())

def compareEdit(request,id=None,mode=None,cNumber=None):
    if mode == "load":  # 由 index.html 按 編輯二 鈕
        unit = Compare.objects.get(id = id)  #取得要修改的資料記
        return render(request, "compareEdit.html", locals())
    elif mode == "save": # 由 edit2.html 按 submit
        unit = Compare.objects.get(id = id)  #取得要修改的資料記錄
        # unit.cNumber = request.POST['cNumber']
        unit.ctable1 = request.POST['ctable1']
        unit.ctable2=request.POST['ctable2']
        unit.cConform=request.POST['cConform']
        unit.cPage=request.POST['cPage']
        unit.cBrand=request.POST['cBrand']
        unit.cRemark=request.POST['cRemark']
        unit.save()  #寫入資料庫
          #寫入資料庫
        message = '已修改...'
        return redirect('/compareIndex/'+str(cNumber)+"/")
    
def tablespecView(request,cNumber=None):
    OrderSpec = Spec.objects.get(cNumber=cNumber)
    Tablespec = tablespec.objects.get(cNumber=cNumber)
    text = ""
    public = {'外板厚度': '0.5t', '外板材質': '鍍鋅板', '內板厚度': '0.5t', '內板材質': '鍍鋅板', '含有風門': '含', 
               '風門型式': '八字開', '馬達絕緣等級': 'F級', '馬達廠牌': '東元', '含有冰水盤管': '含', '鰭片': '一般白鋁', 
               '鰭片厚度': '0.12t', '端側板': '鍍鋅 1.6t', '集水管': '鍍鋅管', '銅管管徑': '1/2"', '銅管厚度': '0.41t', 
               '排水盤厚度': '1.2t', '排水盤材質': '不鏽鋼(304)'}
    new_version = {'外板厚度': Tablespec.first_outthick, '外板材質': Tablespec.first_outmaterial, '內板厚度': Tablespec.first_inthick, 
                   '內板材質': Tablespec.first_inmaterial, '含有風門': Tablespec.six_winddoor, '風門型式': Tablespec.six_form, 
                   '馬達絕緣等級': Tablespec.ten_motor_Insulation, '馬達廠牌': Tablespec.ten_motor_brand, '含有冰水盤管': Tablespec.seven_cold, 
                   '鰭片': Tablespec.seven_fins, '鰭片厚度': Tablespec.seven_thick, '端側板': Tablespec.seven_board, '集水管': Tablespec.seven_waterpipe, 
                   '銅管管徑': Tablespec.seven_size, '銅管厚度': Tablespec.seven_pipe_thick, '排水盤厚度': Tablespec.eight_thick, '排水盤材質': Tablespec.eight_material}
    
    for key in public:
        if key in new_version and public[key] != new_version[key]:
            text += f"{key} => {new_version[key]}\n"

    return render(request, "tablespecView.html",locals())

def tablespecEdit(request,mode=None,cNumber=None):
    OrderSpec = Spec.objects.get(cNumber=cNumber)
    TablespecForm = tablespecForm(request.POST)
    if mode == "load":  # 由 index.html 按 編輯二 鈕
        Tablespec = tablespec.objects.get(cNumber = cNumber)  #取得要修改的資料記
        return render(request, "tablespecEdit.html", locals())
    elif mode == "save": # 由 edit2.html 按 submit
        Tablespec = tablespec.objects.get(cNumber = cNumber)  #取得要修改的資料記錄
        Tablespec.machinename=request.POST['machinename']
        Tablespec.first_boxmaterial=request.POST['first_boxmaterial']
        Tablespec.first_inthick=request.POST['first_inthick']
        Tablespec.first_inmaterial=request.POST['first_inmaterial']
        Tablespec.first_inother=request.POST['first_inother']
        Tablespec.first_outmaterial=request.POST['first_outmaterial']
        Tablespec.first_outthick=request.POST['first_outthick']
        Tablespec.first_boxthick=request.POST['first_boxthick']
        Tablespec.second_door=request.POST['second_door']
        Tablespec.second_frame=request.POST['second_frame']
        Tablespec.third_basic=request.POST['third_basic']
        Tablespec.third_basic_install=request.POST['third_basic_install']
        Tablespec.third_basic_thick=request.POST['third_basic_thick']
        Tablespec.third_basic_brand=request.POST['third_basic_brand']
        Tablespec.third_basic_efficient=request.POST['third_basic_efficient']
        Tablespec.third_basic_frame=request.POST['third_basic_frame']
        Tablespec.third_basic_material=request.POST['third_basic_material']
        Tablespec.third_medium=request.POST['third_medium']
        Tablespec.third_medium_frameform=request.POST['third_medium_frameform']
        Tablespec.third_medium_long=request.POST['third_medium_long']
        Tablespec.third_medium_brand=request.POST['third_medium_brand']
        Tablespec.third_medium_efficient=request.POST['third_medium_efficient']
        Tablespec.third_medium_frame=request.POST['third_medium_frame']
        Tablespec.third_high=request.POST['third_high']
        Tablespec.third_high_frameform=request.POST['third_high_frameform']
        Tablespec.third_high_long=request.POST['third_high_long']
        Tablespec.third_high_brand=request.POST['third_high_brand']
        Tablespec.third_high_efficient=request.POST['third_high_efficient']
        Tablespec.third_high_frame=request.POST['third_high_frame']
        Tablespec.third_basic_medium_material=request.POST['third_basic_medium_material']
        Tablespec.third_high_material=request.POST['third_high_material']
        Tablespec.four_heater=request.POST['four_heater']
        Tablespec.four_material=request.POST['four_material']
        Tablespec.four_way=request.POST['four_way']
        Tablespec.four_SSCR=request.POST['four_SSCR']
        Tablespec.fiv_humidifier=request.POST['fiv_humidifier']
        Tablespec.form=request.POST['fiv_form']
        Tablespec.six_winddoor=request.POST['six_winddoor']
        Tablespec.six_form=request.POST['six_form']
        Tablespec.six_frame_material=request.POST['six_frame_material']
        Tablespec.six_blade_material=request.POST['six_blade_material']
        Tablespec.six_axis=request.POST['six_axis']
        Tablespec.seven_cold=request.POST['seven_cold']
        Tablespec.seven_fins=request.POST['seven_fins']
        Tablespec.seven_thick=request.POST['seven_thick']
        Tablespec.seven_board=request.POST['seven_board']
        Tablespec.seven_pipe_thick=request.POST['seven_pipe_thick']
        Tablespec.seven_waterpipe=request.POST['seven_waterpipe']
        Tablespec.seven_size=request.POST['seven_size']
        Tablespec.seven_hot=request.POST['seven_hot']
        Tablespec.seven_hotfins=request.POST['seven_hotfins']
        Tablespec.seven_hotthick=request.POST['seven_hotthick']
        Tablespec.seven_hotboard=request.POST['seven_hotboard']
        Tablespec.seven_hotpipe_thick=request.POST['seven_hotpipe_thick']
        Tablespec.seven_hotwaterpipe=request.POST['seven_hotwaterpipe']
        Tablespec.seven_hotsize=request.POST['seven_hotsize']
        Tablespec.seven_steam=request.POST['seven_steam']
        Tablespec.seven_steamfins=request.POST['seven_steamfins']
        Tablespec.seven_steamthick=request.POST['seven_steamthick']
        Tablespec.seven_steamboard=request.POST['seven_steamboard']
        Tablespec.seven_steampipe_thick=request.POST['seven_steampipe_thick']
        Tablespec.seven_steamwaterpipe=request.POST['seven_steamwaterpipe']
        Tablespec.seven_steamsize=request.POST['seven_steamsize']
        Tablespec.seven_brine=request.POST['seven_brine']
        Tablespec.seven_brinefins=request.POST['seven_brinefins']
        Tablespec.seven_brinethick=request.POST['seven_brinethick']
        Tablespec.seven_brineboard=request.POST['seven_brineboard']
        Tablespec.seven_brinepipe_thick=request.POST['seven_brinepipe_thick']
        Tablespec.seven_brinewaterpipe=request.POST['seven_brinewaterpipe']
        Tablespec.seven_brinesize=request.POST['seven_brinesize']
        Tablespec.seven_other=request.POST['seven_other']
        Tablespec.eight_material=request.POST['eight_material']
        Tablespec.eight_thick=request.POST['eight_thick']
        Tablespec.nine_brand=request.POST['nine_brand']
        Tablespec.nine_way=request.POST['nine_way']
        Tablespec.nine_form=request.POST['nine_form']
        Tablespec.ten_motor_brand=request.POST['ten_motor_brand']
        Tablespec.ten_motor_elec=request.POST['ten_motor_elec']
        Tablespec.ten_motor_efficient=request.POST['ten_motor_efficient']
        Tablespec.ten_motor_level=request.POST['ten_motor_level']
        Tablespec.ten_motor_open=request.POST['ten_motor_open']
        Tablespec.ten_motor_cold=request.POST['ten_motor_cold']
        Tablespec.ten_motor_ask=request.POST['ten_motor_ask']
        Tablespec.ten_motor_form=request.POST['ten_motor_form']
        Tablespec.ten_motor_Insulation=request.POST['ten_motor_form']
        Tablespec.eleven_material=request.POST['eleven_material']
        Tablespec.eleven_form=request.POST['eleven_form']
        Tablespec.eleven_key=request.POST['eleven_key']
        Tablespec.eleven_level=request.POST['eleven_level']
        Tablespec.eleven_size=request.POST['eleven_size']
        Tablespec.twelve_view=request.POST['twelve_view']
        Tablespec.twelve_pressure=request.POST['twelve_pressure']
        Tablespec.twelve_unit=request.POST['twelve_unit']
        Tablespec.twelve_brand=request.POST['twelve_brand']
        Tablespec.twelve_light=request.POST['twelve_light']
        Tablespec.twelve_level=request.POST['twelve_level']
        Tablespec.twelve_light_elec=request.POST['twelve_light_elec']
        Tablespec.twelve_light_open=request.POST['twelve_light_open']
        Tablespec.twelve_winddoor=request.POST['twelve_winddoor']
        Tablespec.twelve_net=request.POST['twelve_net']
        Tablespec.twelve_material=request.POST['twelve_material']
        Tablespec.twelve_silicone=request.POST['twelve_silicone']
        Tablespec.twelve_roof=request.POST['twelve_roof']
        Tablespec.twelve_newmaterial=request.POST['twelve_newmaterial']
        Tablespec.twelve_location=request.POST['twelve_location']
        Tablespec.twelve_test=request.POST['twelve_test']
        Tablespec.thirteen_way=request.POST['thirteen_way']
        Tablespec.thirteen_elevator=request.POST['thirteen_elevator']
        Tablespec.thirteen_box=request.POST['thirteen_box']
        Tablespec.thirteen_into=request.POST['thirteen_into']
        Tablespec.thirteen_shock=request.POST['thirteen_shock']
        Tablespec.thirteen_shocbox=request.POST['thirteen_shocbox']
        Tablespec.thirteen_apply=request.POST['thirteen_apply']
        Tablespec.thirteen_fix=request.POST['thirteen_fix']
        Tablespec.thirteen_budget=request.POST['thirteen_budget']
        Tablespec.thirteen_displace=request.POST['thirteen_displace']
        Tablespec.thirteen_fixway=request.POST['thirteen_fixway']
        Tablespec.fourteen_ass=request.POST['fourteen_ass']
        Tablespec.fourteen_budget=request.POST['fourteen_budget']
        Tablespec.fourteen_time=request.POST['fourteen_time']
        Tablespec.fourteen_into=request.POST['fourteen_into']
        Tablespec.fourteen_intotime=request.POST['fourteen_intotime']
        Tablespec.fourteen_afterass=request.POST['fourteen_afterass']
        Tablespec.fourteen_afterasstime=request.POST['fourteen_afterasstime']
        Tablespec.fourteen_foreigninto=request.POST['fourteen_foreigninto']
        Tablespec.fourteen_foreignintotime=request.POST['fourteen_foreignintotime']
        Tablespec.fourteen_card=request.POST['fourteen_card']
        Tablespec.fourteen_insurance=request.POST['fourteen_insurance']
        Tablespec.fifteen_location=request.POST['fifteen_location']
        Tablespec.fifteen_level=request.POST['fifteen_level']
        Tablespec.fifteen_vertical=request.POST['fifteen_vertical']
        Tablespec.fifteen_box=request.POST['fifteen_box']
        Tablespec.sixteen_install=request.POST['sixteen_install']
        Tablespec.sixteen_prepare=request.POST['sixteen_prepare']
        Tablespec.seventeen_silicone=request.POST['seventeen_silicone']
        Tablespec.seventeen_brand=request.POST['seventeen_brand']
        Tablespec.eighteen_other=request.POST['eighteen_other']
        Tablespec.save()  #寫入資料庫
          #寫入資料庫
        return redirect('/tablespecView/'+str(cNumber)+'/')

def tablespecPost(request,cNumber=None):
    text = ""
    CompareSpec = Spec.objects.get(cNumber=cNumber)
    
    brands = ["高雄日月光","中壢日月光","美光","長春","旭富","天基","同步輻射","南亞電路","力成","華新科","漢欣","金像電子","喬輝","大立光","洋基","瑞昱","迅展","御勤","新竹南茂"]
    special = "123"
    for brand in brands:
        if brand in CompareSpec.cClient or brand in CompareSpec.cProject:
            print(brand)
            try:
                special = Special.objects.get(cClient=brand)
            except:
                special = "123"
    temp = {'machinename':' ',
            'first_boxmaterial':' ','first_inthick':'','first_inmaterial':'','first_inother':'','first_outmaterial':'','first_outthick':'','first_boxthick':'',
            'second_door':'','second_frame':'',
            'third_basic':'','third_basic_install':'','third_basic_thick':'','third_basic_brand':'','third_basic_efficient':'','third_basic_frame':'','third_basic_material':'','third_medium':'','third_medium_frameform':'','third_medium_long':'','third_medium_brand':'','third_medium_efficient':'','third_medium_frame':'',
            'third_high':'','third_high_frameform':'','third_high_long':'','third_high_brand':'','third_high_efficient':'','third_high_frame':'','third_basic_medium_material':'','third_high_material':'',
            'four_heater':'','four_material':'','four_way':'','four_SSCR':'',
            'fiv_humidifier':'','fiv_form':'',
            'six_winddoor':'','six_form':'','six_frame_material':'','six_blade_material':'','six_axis':'',
            'seven_cold':'','seven_fins':'','seven_thick':'','seven_board':'','seven_pipe_thick':'','seven_waterpipe':'','seven_size':'','seven_hot':'','seven_hotfins':'','seven_hotthick':'','seven_hotboard':'','seven_hotpipe_thick':'','seven_hotwaterpipe':'','seven_hotsize':'','seven_other':'',
            'eight_material':'','eight_thick':'','seven_steam':'','seven_steamfins':'','seven_steamthick':'','seven_steamboard':'','seven_steampipe_thick':'','seven_steamwaterpipe':'','seven_steamsize':'','seven_brine':'','seven_brinefins':'','seven_brinethick':'','seven_brineboard':'','seven_brinepipe_thick':'','seven_brinewaterpipe':'','seven_brinesize':'',
            'nine_brand':'科祿格','nine_way':'','nine_form':'',
            'ten_motor_brand':'東元','ten_motor_elec':'接至端子台','ten_motor_efficient':'','ten_motor_level':'','ten_motor_open':'一般','ten_motor_cold':'自立通風','ten_motor_ask':'','ten_motor_form':'','ten_motor_Insulation':'',
            'eleven_material':'','eleven_form':'','eleven_key':'','eleven_level':'','eleven_size':'',
            'twelve_view':'','twelve_pressure':'','twelve_unit':'','twelve_brand':'','twelve_light':'','twelve_level':'','twelve_light_elec':'各箱段獨立電源','twelve_light_open':'','twelve_winddoor':'','twelve_net':'','televe_material':'','twelve_silicone':'勝新標準','twelve_location':'','twelve_roof':'','twelve_newmaterial':'','twelve_test':'',
            'thirteen_way':'','thirteen_elevator':'','thirteen_box':'','thirteen_into':'','thirteen_shock':'','thirteen_shocbox':'','thirteen_apply':'','thirteen_fix':'','thirteen_budget':'','thirteen_budgetdollar':'','thirteen_displace':'無','thirteen_fixway':'無',
            'fourteen_ass':'','fourteen_budget':'','fourteen_budgetdollar':'','fourteen_time':'','fourteen_into':'','fourteen_afterass':'','fourteen_foreigninto':'','fourteen_card':'','fourteen_insurance':'',
            'fifteen_location':'','fifteen_level':'','fifteen_vertical':'','fifteen_box':'',
            'sixteen_install':'','sixteen_prepare':'',
            'seventeen_silicone':'勝新標準','seventeen_brand':'一般',
            'eighteen_other':''}
    
    aaa = Compare.objects.filter(ref=CompareSpec)
    pdf = pdfplumber.open("./media/"+str(CompareSpec.cFile))
    for page in pdf.pages:
        page_text = page.extract_text()
        page_text = re.sub(r'\d+ \|', "", page_text)
        page_text = re.sub(r'[\d]* of [\d]*',"",page_text)
        page_text = re.sub(r'第 \d+ 章\n空氣調節箱|第 \d+ 章\n空氣調節箱', "", page_text)
        text += page_text   
    text = text.replace(' ', '').replace('\n',"").replace("(","").replace(")","").replace("]","").replace("[","")
    answer = retext(text)
    temp.update(answer)
    if request.method == 'POST':
        tablespecform = tablespecForm(request.POST)
        if tablespecform.is_valid():
            cNumber =  cNumber
            machinename=tablespecform.cleaned_data['machinename']
            first_boxmaterial=tablespecform.cleaned_data['first_boxmaterial']
            first_inthick=tablespecform.cleaned_data['first_inthick']
            first_inmaterial=tablespecform.cleaned_data['first_inmaterial']
            first_inother=tablespecform.cleaned_data['first_inother']
            first_outmaterial=tablespecform.cleaned_data['first_outmaterial']
            first_outthick=tablespecform.cleaned_data['first_outthick']
            first_boxthick=tablespecform.cleaned_data['first_boxthick']
            second_door=tablespecform.cleaned_data['second_door']
            second_frame=tablespecform.cleaned_data['second_frame']
            third_basic=tablespecform.cleaned_data['third_basic']
            third_basic_install=tablespecform.cleaned_data['third_basic_install']
            third_basic_thick=tablespecform.cleaned_data['third_basic_thick']
            third_basic_brand=tablespecform.cleaned_data['third_basic_brand']
            third_basic_efficient=tablespecform.cleaned_data['third_basic_efficient']
            third_basic_frame=tablespecform.cleaned_data['third_basic_frame']
            third_basic_material=tablespecform.cleaned_data['third_basic_material']
            third_medium=tablespecform.cleaned_data['third_medium']
            third_medium_frameform=tablespecform.cleaned_data['third_medium_frameform']
            third_medium_long=tablespecform.cleaned_data['third_medium_long']
            third_medium_brand=tablespecform.cleaned_data['third_medium_brand']
            third_medium_efficient=tablespecform.cleaned_data['third_medium_efficient']
            third_medium_frame=tablespecform.cleaned_data['third_medium_frame']
            third_high=tablespecform.cleaned_data['third_high']
            third_high_frameform=tablespecform.cleaned_data['third_high_frameform']
            third_high_long=tablespecform.cleaned_data['third_high_long']
            third_high_brand=tablespecform.cleaned_data['third_high_brand']
            third_high_efficient=tablespecform.cleaned_data['third_high_efficient']
            third_high_frame=tablespecform.cleaned_data['third_high_frame']
            third_basic_medium_material=tablespecform.cleaned_data['third_basic_medium_material']
            third_high_material=tablespecform.cleaned_data['third_high_material']
            four_heater=tablespecform.cleaned_data['four_heater']
            four_material=tablespecform.cleaned_data['four_material']
            four_way=tablespecform.cleaned_data['four_way']
            four_SSCR=tablespecform.cleaned_data['four_SSCR']
            fiv_humidifier=tablespecform.cleaned_data['fiv_humidifier']
            fiv_form=tablespecform.cleaned_data['fiv_form']
            six_winddoor=tablespecform.cleaned_data['six_winddoor']
            six_form=tablespecform.cleaned_data['six_form']
            six_frame_material=tablespecform.cleaned_data['six_frame_material']
            six_blade_material=tablespecform.cleaned_data['six_blade_material']
            six_axis=tablespecform.cleaned_data['six_axis']
            seven_cold=tablespecform.cleaned_data['seven_cold']
            seven_fins=tablespecform.cleaned_data['seven_fins']
            seven_thick=tablespecform.cleaned_data['seven_thick']
            seven_board=tablespecform.cleaned_data['seven_board']
            seven_pipe_thick=tablespecform.cleaned_data['seven_pipe_thick']
            seven_waterpipe=tablespecform.cleaned_data['seven_waterpipe']
            seven_size=tablespecform.cleaned_data['seven_size']
            seven_hot=tablespecform.cleaned_data['seven_hot']
            seven_hotfins=tablespecform.cleaned_data['seven_hotfins']
            seven_hotthick=tablespecform.cleaned_data['seven_hotthick']
            seven_hotboard=tablespecform.cleaned_data['seven_hotboard']
            seven_hotpipe_thick=tablespecform.cleaned_data['seven_hotpipe_thick']
            seven_hotwaterpipe=tablespecform.cleaned_data['seven_hotwaterpipe']
            seven_hotsize=tablespecform.cleaned_data['seven_hotsize']
            seven_steam=tablespecform.cleaned_data['seven_steam']
            seven_steamfins=tablespecform.cleaned_data['seven_steamfins']
            seven_steamthick=tablespecform.cleaned_data['seven_steamthick']
            seven_steamboard=tablespecform.cleaned_data['seven_steamboard']
            seven_steampipe_thick=tablespecform.cleaned_data['seven_steampipe_thick']
            seven_steamwaterpipe=tablespecform.cleaned_data['seven_steamwaterpipe']
            seven_steamsize=tablespecform.cleaned_data['seven_steamsize']
            seven_brine=tablespecform.cleaned_data['seven_brine']
            seven_brinefins=tablespecform.cleaned_data['seven_brinefins']
            seven_brinethick=tablespecform.cleaned_data['seven_brinethick']
            seven_brineboard=tablespecform.cleaned_data['seven_brineboard']
            seven_brinepipe_thick=tablespecform.cleaned_data['seven_brinepipe_thick']
            seven_brinewaterpipe=tablespecform.cleaned_data['seven_brinewaterpipe']
            seven_brinesize=tablespecform.cleaned_data['seven_brinesize']
            seven_other=tablespecform.cleaned_data['seven_other']
            eight_material=tablespecform.cleaned_data['eight_material']
            eight_thick=tablespecform.cleaned_data['eight_thick']
            nine_brand=tablespecform.cleaned_data['nine_brand']
            nine_way=tablespecform.cleaned_data['nine_way']
            nine_form=tablespecform.cleaned_data['nine_form']
            ten_motor_brand=tablespecform.cleaned_data['ten_motor_brand']
            ten_motor_elec=tablespecform.cleaned_data['ten_motor_elec']
            ten_motor_efficient=tablespecform.cleaned_data['ten_motor_efficient']
            ten_motor_level=tablespecform.cleaned_data['ten_motor_level']
            ten_motor_open=tablespecform.cleaned_data['ten_motor_open']
            ten_motor_cold=tablespecform.cleaned_data['ten_motor_cold']
            ten_motor_ask=tablespecform.cleaned_data['ten_motor_ask']
            ten_motor_form=tablespecform.cleaned_data['ten_motor_form']
            ten_motor_Insulation=tablespecform.cleaned_data['ten_motor_Insulation']
            eleven_material=tablespecform.cleaned_data['eleven_material']
            eleven_form=tablespecform.cleaned_data['eleven_form']
            eleven_key=tablespecform.cleaned_data['eleven_key']
            eleven_level=tablespecform.cleaned_data['eleven_level']
            eleven_size=tablespecform.cleaned_data['eleven_size']
            twelve_view=tablespecform.cleaned_data['twelve_view']
            twelve_pressure=tablespecform.cleaned_data['twelve_pressure']
            twelve_unit=tablespecform.cleaned_data['twelve_unit']
            twelve_brand=tablespecform.cleaned_data['twelve_brand']
            twelve_light=tablespecform.cleaned_data['twelve_light']
            twelve_level=tablespecform.cleaned_data['twelve_level']
            twelve_light_elec=tablespecform.cleaned_data['twelve_light_elec']
            twelve_light_open=tablespecform.cleaned_data['twelve_light_open']
            twelve_winddoor=tablespecform.cleaned_data['twelve_winddoor']
            twelve_net=tablespecform.cleaned_data['twelve_net']
            twelve_material=tablespecform.cleaned_data['twelve_material']
            twelve_silicone=tablespecform.cleaned_data['twelve_silicone']
            twelve_location=tablespecform.cleaned_data['twelve_location']
            twelve_roof=tablespecform.cleaned_data['twelve_roof']
            twelve_newmaterial=tablespecform.cleaned_data['twelve_newmaterial']
            twelve_test=tablespecform.cleaned_data['twelve_test']
            thirteen_way=tablespecform.cleaned_data['thirteen_way']
            thirteen_elevator=tablespecform.cleaned_data['thirteen_elevator']
            thirteen_box=tablespecform.cleaned_data['thirteen_box']
            thirteen_into=tablespecform.cleaned_data['thirteen_into']
            thirteen_shock=tablespecform.cleaned_data['thirteen_shock']
            thirteen_shocbox=tablespecform.cleaned_data['thirteen_shocbox']
            thirteen_apply=tablespecform.cleaned_data['thirteen_apply']
            thirteen_fix=tablespecform.cleaned_data['thirteen_fix']
            thirteen_budget=tablespecform.cleaned_data['thirteen_budget']
            thirteen_budgetdollar=tablespecform.cleaned_data['thirteen_budgetdollar']
            thirteen_displace=tablespecform.cleaned_data['thirteen_displace']
            thirteen_fixway=tablespecform.cleaned_data['thirteen_fixway']
            fourteen_ass=tablespecform.cleaned_data['fourteen_ass']
            fourteen_budget=tablespecform.cleaned_data['fourteen_budget']
            fourteen_budgetdollar=tablespecform.cleaned_data['fourteen_budgetdollar']
            fourteen_time=tablespecform.cleaned_data['fourteen_time']
            fourteen_into=tablespecform.cleaned_data['fourteen_into']
            fourteen_intotime=tablespecform.cleaned_data['fourteen_intotime']
            fourteen_afterass=tablespecform.cleaned_data['fourteen_afterass']
            fourteen_afterasstime=tablespecform.cleaned_data['fourteen_afterasstime']
            fourteen_foreigninto=tablespecform.cleaned_data['fourteen_foreigninto']
            fourteen_foreignintotime=tablespecform.cleaned_data['fourteen_foreignintotime']
            fourteen_card=tablespecform.cleaned_data['fourteen_card']
            fourteen_insurance=tablespecform.cleaned_data['fourteen_insurance']
            fifteen_location=tablespecform.cleaned_data['fifteen_location']
            fifteen_level=tablespecform.cleaned_data['fifteen_level']
            fifteen_vertical=tablespecform.cleaned_data['fifteen_vertical']
            fifteen_box=tablespecform.cleaned_data['fifteen_box']
            sixteen_install=tablespecform.cleaned_data['sixteen_install']
            sixteen_prepare=tablespecform.cleaned_data['sixteen_prepare']
            seventeen_silicone=tablespecform.cleaned_data['seventeen_silicone']
            seventeen_brand=tablespecform.cleaned_data['seventeen_brand']
            eighteen_other=tablespecform.cleaned_data['eighteen_other']


            tablespecUnit = tablespec.objects.create(home=CompareSpec, cNumber=cNumber, 
                        machinename=machinename,
                        first_boxmaterial=first_boxmaterial,
                        first_inthick=first_inthick,
                        first_inmaterial=first_inmaterial,
                        first_inother=first_inother,
                        first_outmaterial=first_outmaterial,
                        first_outthick=first_outthick,
                        first_boxthick=first_boxthick,
                        second_door=second_door,
                        second_frame=second_frame,
                        third_basic=third_basic,
                        third_basic_install=third_basic_install,
                        third_basic_thick=third_basic_thick,
                        third_basic_brand=third_basic_brand,
                        third_basic_efficient=third_basic_efficient,
                        third_basic_frame=third_basic_frame,
                        third_basic_material=third_basic_material,
                        third_medium=third_medium,
                        third_medium_frameform=third_medium_frameform,
                        third_medium_long=third_medium_long,
                        third_medium_brand=third_medium_brand,
                        third_medium_efficient=third_medium_efficient,
                        third_medium_frame=third_medium_frame,
                        third_high=third_high,
                        third_high_frameform=third_high_frameform,
                        third_high_long=third_high_long,
                        third_high_brand=third_high_brand,
                        third_high_efficient=third_high_efficient,
                        third_high_frame=third_high_frame,
                        third_high_material=third_high_material,
                        third_basic_medium_material=third_basic_medium_material,
                        four_heater=four_heater,
                        four_material=four_material,
                        four_way=four_way,
                        four_SSCR=four_SSCR,
                        fiv_humidifier=fiv_humidifier,
                        fiv_form=fiv_form,
                        six_winddoor=six_winddoor,
                        six_form=six_form,
                        six_frame_material=six_frame_material,
                        six_blade_material=six_blade_material,
                        six_axis=six_axis,
                        seven_cold=seven_cold,
                        seven_fins=seven_fins,
                        seven_thick=seven_thick,
                        seven_board=seven_board,
                        seven_pipe_thick=seven_pipe_thick,
                        seven_waterpipe=seven_waterpipe,
                        seven_size=seven_size,
                        seven_other=seven_other,
                        seven_hot=seven_hot,
                        seven_hotfins=seven_hotfins,
                        seven_hotthick=seven_hotthick,
                        seven_hotboard=seven_hotboard,
                        seven_hotpipe_thick=seven_hotpipe_thick,
                        seven_hotwaterpipe=seven_hotwaterpipe,
                        seven_hotsize=seven_hotsize,
                        seven_steam=seven_steam,
                        seven_steamfins=seven_steamfins,
                        seven_steamthick=seven_steamthick,
                        seven_steamboard=seven_steamboard,
                        seven_steampipe_thick=seven_steampipe_thick,
                        seven_steamwaterpipe=seven_steamwaterpipe,
                        seven_steamsize=seven_steamsize,
                        seven_brine=seven_brine,
                        seven_brinefins=seven_brinefins,
                        seven_brinethick=seven_brinethick,
                        seven_brineboard=seven_brineboard,
                        seven_brinepipe_thick=seven_brinepipe_thick,
                        seven_brinewaterpipe=seven_brinewaterpipe,
                        seven_brinesize=seven_brinesize,
                        eight_material=eight_material,
                        eight_thick=eight_thick,
                        nine_brand=nine_brand,
                        nine_way=nine_way,
                        nine_form=nine_form,
                        ten_motor_brand=ten_motor_brand,
                        ten_motor_elec=ten_motor_elec,
                        ten_motor_efficient=ten_motor_efficient,
                        ten_motor_level=ten_motor_level,
                        ten_motor_open=ten_motor_open,
                        ten_motor_cold=ten_motor_cold,
                        ten_motor_ask=ten_motor_ask,
                        ten_motor_form=ten_motor_form,
                        ten_motor_Insulation=ten_motor_Insulation,
                        eleven_material=eleven_material,
                        eleven_form=eleven_form,
                        eleven_key=eleven_key,
                        eleven_level=eleven_level,
                        eleven_size=eleven_size,
                        twelve_view=twelve_view,
                        twelve_pressure=twelve_pressure,
                        twelve_unit=twelve_unit,
                        twelve_brand=twelve_brand,
                        twelve_light=twelve_light,
                        twelve_level=twelve_level,
                        twelve_light_elec=twelve_light_elec,
                        twelve_light_open=twelve_light_open,
                        twelve_winddoor=twelve_winddoor,
                        twelve_net=twelve_net,
                        twelve_material=twelve_material,
                        twelve_silicone=twelve_silicone,
                        twelve_location=twelve_location,
                        twelve_roof=twelve_roof,
                        twelve_newmaterial=twelve_newmaterial,
                        twelve_test=twelve_test,
                        thirteen_way=thirteen_way,
                        thirteen_elevator=thirteen_elevator,
                        thirteen_box=thirteen_box,
                        thirteen_into=thirteen_into,
                        thirteen_shock=thirteen_shock,
                        thirteen_shocbox=thirteen_shocbox,
                        thirteen_apply=thirteen_apply,
                        thirteen_fix=thirteen_fix,
                        thirteen_budget=thirteen_budget,
                        thirteen_budgetdollar=thirteen_budgetdollar,
                        thirteen_displace=thirteen_displace,
                        thirteen_fixway=thirteen_fixway,
                        fourteen_ass=fourteen_ass,
                        fourteen_budget=fourteen_budget,
                        fourteen_budgetdollar=fourteen_budgetdollar,
                        fourteen_time=fourteen_time,
                        fourteen_into=fourteen_into,
                        fourteen_intotime=fourteen_intotime,
                        fourteen_afterass=fourteen_afterass,
                        fourteen_afterasstime=fourteen_afterasstime,
                        fourteen_foreigninto=fourteen_foreigninto,
                        fourteen_foreignintotime=fourteen_foreignintotime,
                        fourteen_card=fourteen_card,
                        fourteen_insurance=fourteen_insurance,
                        fifteen_location=fifteen_location,
                        fifteen_level=fifteen_level,
                        fifteen_vertical=fifteen_vertical,
                        fifteen_box=fifteen_box,
                        sixteen_install=sixteen_install,
                        sixteen_prepare=sixteen_prepare,
                        seventeen_silicone=seventeen_silicone,
                        seventeen_brand=seventeen_brand,
                        eighteen_other=eighteen_other)
            tablespecUnit.save()
            CompareSpec.chaveorder="有"
            CompareSpec.save()
            return redirect('/tablespecView/'+str(cNumber)+'/')
        else:
             message="驗證錯誤"
    else:
         message=''
    tablespecform = tablespecForm(temp)
    return render(request, "tablespecPost.html",locals())

def retext(text):
    first_material_select = {"鍍鋅":"鍍鋅板","烤漆":"烤漆板","鹽化":"鹽化鋼板","PVC覆膜彩色":"鍍鋅板","彩色烤漆":"烤漆板","不銹":"SUS304","不鏽":"SUS304"}
    sixform_select = {"八字形":"八字開","平形":"平形開"}
    seven_pipe = {"鍍鋅鋼管":"鍍鋅管","不鏽鋼管":"不鏽鋼管"}
    dist = {}
    # temp = re.findall('內外層鋼板[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     if result[0] == "1":
    #         dist["first_outthick"]=str(1.0)+'t'
    #         dist["first_inthick"]=str(1.0)+'t'
    #     else:
    #         dist["first_outthick"]=result[0]+'t'
    #         dist["first_inthick"]=result[0]+'t'
    #     dist["first_outmaterial"]=first_material_select[result[1]]
    #     dist["first_inmaterial"]=first_material_select[result[1]]
    # temp = re.findall('外層鋼板+[\u4e00-\u9fa5]*(\d+\.\d+)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     dist["first_outthick"]=result[0]+'t'
    #     dist["first_outmaterial"]=first_material_select[result[1]]
    #     #配電箱和外板材質相同(不確定PVC是甚麼材質)
    # temp = re.findall('加強板組裝而成,(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     dist["first_outthick"]=result[0]+'t'
    #     dist["first_outmaterial"]=first_material_select[result[1]]
    # temp = re.findall('內層鋼板+[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     dist["first_inthick"]=result[0]+'t'
    #     dist["first_inmaterial"]=first_material_select[result[1]]
    # temp = re.findall('空氣混合箱之進風口需配置([\w]*)風門',text)
    # if(len(temp) != 0):
    #     dist["six_winddoor"]="含"
    #     dist["six_form"]=sixform_select[temp[0]]
    # temp = re.findall('馬達採([A-Za-z]級)([\w]*)馬達',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     dist["ten_motor_Insulation"]=result[0]
    #     dist["ten_motor_brand"]=result[1]
    # if "冰水盤管" in text:
    #     dist["seven_cold"] = "含"
    #     temp = re.findall('鋁[\w]*鰭片[\D]*(\d.\d\d|\d.\d)[mm|mmt]*',text)
    #     if(len(temp) != 0):
    #         dist["seven_fins"]="一般白鋁"
    #         dist["seven_thick"]=temp[0]+'t'
    #     temp = re.findall('端側板[\D]*(\d.\d)[mm|mmt]*以上([\w]*)鋼板',text)
    #     if(len(temp) != 0):
    #         result = list(temp[0])
    #         if result[1] == "鍍鋅":
    #             dist["seven_board"]="鍍鋅"+' 1.6t'
    #         if result[1] == "不鏽鋼":
    #             dist["seven_board"]="不鏽鋼"+' 1.5t'
    #     temp = re.findall('上、下、左、右蓋板[\w]*為([\w]*)材質',text)
    #     if(len(temp) != 0):
    #         result = list(temp[0])
    #         result = [''.join(result)]
    #         if result[0] == "鍍鋅":
    #             dist["seven_board"]= "鍍鋅"+' 1.6t'
    #         if result[0] == "不鏽鋼" or result[0] == "不銹鋼":
    #             dist["seven_board"]= "不鏽鋼"+' 1.5t'
    #     temp = re.findall('集流管[\u4e00-\u9fa5]*為([\w]*鋼管)',text)
    #     if(len(temp) != 0):
    #         dist["seven_waterpipe"]=seven_pipe[temp[0]]
    #     temp = re.findall('盤管為(\d+/\d+\")OD[\D]*(\d.\d\d|\d.\d|\d)[mm|mmt]+[\u4e00-\u9fa5]*',text)
    #     if(len(temp) != 0):
    #         result = list(temp[0])
    #         dist["seven_size"]=result[0]
    #         dist["seven_pipe_thick"] = result[1]+'t'
    # temp = re.findall('採用(\d.\d)[tmm厚度|mmt厚度|mm厚度]*([\w]*鋼板)製作',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     dist["eight_thick"]=result[0]+'t'
    #     if "SUS304" in result[1]:
    #         dist["eight_material"]="不鏽鋼(304)"
    # temp = re.findall('厚度規號為([\d]*)[\w]*不鏽鋼',text)
    # if(len(temp) != 0):
    #     result = list(temp[0])
    #     result = [''.join(result)]
    #     if "16" in result[0]:
    #         dist["eight_material"]="不鏽鋼(316)"
    #     if "04" in result[0]:
    #         dist["eight_material"]="不鏽鋼(304)"
    #     if "430" in result[0]:
    #         dist["eight_material"]="不鏽鋼(430)"
    temp = re.findall('內外層鋼板[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    if(len(temp) != 0):
        result = list(temp[0])
        if result[0] == "1":
            dist["first_outthick"]=str(1.0)+'t'
            dist["first_inthick"]=str(1.0)+'t'
        else:
            dist["first_outthick"]=result[0]+'t'
            dist["first_inthick"]=result[0]+'t'
        dist["first_outmaterial"]=first_material_select[result[1]]
        dist["first_inmaterial"]=first_material_select[result[1]]
    temp = re.findall('外層板+[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt|mmT]+([\w]*)鋼板',text)
    if(len(temp) != 0):
        result = list(temp[0])
        if result[0] == "1":
            dist["first_outthick"]=str(1.0)+'t'
        else:
            dist["first_outthick"]=result[0]+'t'
        dist["first_outmaterial"]=first_material_select[result[1]]
    temp = re.findall('內層板+[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt|mmT]+([\w]*)鋼板',text)
    if(len(temp) != 0):
        result = list(temp[0])
        if result[0] == "1":
            dist["first_inthick"]=str(1.0)+'t'
        else:
            dist["first_inthick"]=result[0]+'t'
        dist["first_inmaterial"]=first_material_select[result[1]]
    temp = re.findall('外層鋼板+[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    if(len(temp) != 0):
        result = list(temp[0])
        if result[0] == "1":
            dist["first_outthick"]=str(1.0)+'t'
        else:
            dist["first_outthick"]=result[0]+'t'
        dist["first_outmaterial"]=first_material_select[result[1]]
    temp = re.findall('雙面([\w]*)鋼板構造',text)
    if(len(temp) != 0):
        result = list(temp[0])
        result = [''.join(result)]
        dist["first_outmaterial"]=first_material_select[result[0]]
        dist["first_inmaterial"]=first_material_select[result[0]]
    temp = re.findall('加強板組裝而成,(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    if(len(temp) != 0):
        result = list(temp[0])
        if result[0] == "1":
            dist["first_outthick"]=str(1.0)+'t'
        else:
            dist["first_outthick"]=result[0]+'t'
        dist["first_outmaterial"]=first_material_select[result[1]]
    temp = re.findall('內層鋼板+[\u4e00-\u9fa5]*(\d+\.\d+|\d)*[mm|mmt]+[\u4e00-\u9fa5]*之([\w]*)鋼板',text)
    if(len(temp) != 0):
        result = list(temp[0])
        if result[0] == "1":
            dist["first_inthick"]=str(1.0)+'t'
        else:
            dist["first_inthick"]=result[0]+'t'
        dist["first_inmaterial"]=first_material_select[result[1]]
    temp = re.findall('空氣混合箱之進風口需配置([\w]*)風門',text)
    if(len(temp) != 0):
        dist["six_winddoor"]="含"
        dist["six_form"]=sixform_select[temp[0]]
    temp = re.findall('([A-Za-z]級)以上絕緣',text)
    if(len(temp) != 0):
        result = list(temp[0])
        result = [''.join(result)]
        dist["ten_motor_Insulation"]=result[0]
    temp = re.findall('馬達+[\u4e00-\u9fa5]*採用(東元)?(或)?(大同)?([\u4e00-\u9fa5]*)?高效率馬達採用([A-Za-z]級)',text)
    if(len(temp) != 0):
        result = list(temp[0])
        dist["ten_motor_Insulation"]=result[4]
        if(result[0] == "東元"):
            dist["ten_motor_brand"]="東元"
        elif(result[2] == "東元"): 
            dist["ten_motor_brand"]="大同"
        elif(result[0] == "" and result[2] == ""):
            dist["ten_motor_brand"]=result[3]
        if ("高效率" in text):
            dist["ten_motor_efficient"]="高效率"
        elif ("超高效率" in text):
            dist["ten_motor_efficient"]="超高效率"

    temp = re.findall('馬達採([A-Za-z]級)([\w]*)馬達',text)
    if(len(temp) != 0):
        result = list(temp[0])
        dist["ten_motor_Insulation"]=result[0]
        dist["ten_motor_brand"]=result[1]

    if "冰水盤管" in text:
        dist["seven_cold"] = "含"
        temp = re.findall('鋁[\w]*鰭片[\D]*(\d.\d\d|\d.\d)[mm|mmt]*',text)
        if(len(temp) != 0):
            dist["seven_fins"]="一般白鋁"
            dist["seven_thick"]=temp[0]+'t'
        if "鋁鰭片" in text:
            dist["seven_fins"]= "一般白鋁"
        if "鋁鰭片" in text and "藍波" in text:
            dist["seven_fins"]= "藍波鋁片"
        temp = re.findall('端側板[\D]*(\d.\d)[mm|mmt]*以上([\w]*)鋼板',text)
        if(len(temp) != 0):
            result = list(temp[0])
            if result[1] == "鍍鋅":
                dist["seven_board"]="鍍鋅"+' 1.6t'
            if result[1] == "不鏽鋼":
                dist["seven_board"]="不鏽鋼"+' 1.5t'
        temp = re.findall('鰭片厚度不小於t=(\d.\d\d|\d.\d)[mm|mmt]*',text)
        if(len(temp) != 0):
            dist["seven_thick"]=temp[0]+'t'
        temp = re.findall('鰭片[\w]+厚度為\d.[\d]*英吋(\d.\d\d)[mm|mm|mmT]+[以上]+',text)
        if(len(temp) != 0):
            dist["seven_thick"]=temp[0]+'t'
        temp = re.findall('上、下、左、右蓋板[\w]*為([\w]*)材質',text)
        if(len(temp) != 0):
            result = list(temp[0])
            result = [''.join(result)]
            if result[0] == "鍍鋅":
                dist["seven_board"]= "鍍鋅"+' 1.6t'
            if result[0] == "不鏽鋼" or result[0] == "不銹鋼":
                dist["seven_board"]= "不鏽鋼"+' 1.5t'
        temp = re.findall('集流管[\u4e00-\u9fa5]*為([\w]*鋼管)',text)
        if(len(temp) != 0):
            dist["seven_waterpipe"]=seven_pipe[temp[0]]
        temp = re.findall('盤管為(\d+/\d+)[”|\"][\D]*(\d.\d\d|\d.\d|\d)[mm|mmt]+[\u4e00-\u9fa5]*',text)
        if(len(temp) != 0):
            result = list(temp[0])
            dist["seven_size"]=result[0]+"\""
            dist["seven_pipe_thick"] = result[1]+'t'
        temp = re.findall('材質為(\d+/\d+)[”|\"]OD',text)
        if(len(temp) != 0):
            result = list(temp[0])
            result = [''.join(result)]
            dist["seven_size"]=result[0]+"\""
        temp = re.findall('t=(\d.\d\d|\d.\d|\d)[mm|mmt]*去氧無縫紫銅管',text)
        if(len(temp) != 0):
            result = list(temp[0])
            result = [''.join(result)]
            dist["seven_pipe_thick"] = result[0]+'t'
    temp = re.findall('(採用|使用)(\d.\d)[mmt|tmm|mm|tmm厚度|mmt厚度|mm厚度]*[以上]*([\w]*板)(製作|製造)',text)
    if(len(temp) != 0):
        result = list(temp[0])
        dist["eight_thick"]=result[1]+'t'
        if "SUS304" in result[2] or "不銹鋼" in result[2] or "不鏽鋼" in result[2]:
            dist["eight_material"]="不鏽鋼(304)"
        if "鍍鋅" in result[2]:
            dist["eight_material"]="鍍鋅"
    temp = re.findall('(採用|使用)[\w]+(\d.\d)[\w]*#[\d]*[以上]*([\u4e00-\u9fa5]*板)(製作|製造)',text)
    if(len(temp) != 0):
        result = list(temp[0])
        dist["eight_thick"]=result[1]+'t'
        if "SUS304" in result[2] or "不銹鋼" in result[2] or "不鏽鋼" in result[2]:
            dist["eight_material"]="不鏽鋼(304)"
        if "鍍鋅" in result[2]:
            dist["eight_material"]="鍍鋅"
    temp = re.findall('(採用|使用)[\w]*(\d.\d)[mm|mmt|mmT]*(\w*板)[製造|製成|製作]+',text)
    if(len(temp) != 0):
        result = list(temp[0])
        dist["eight_thick"]=result[1]+'t'
        if "SUS304" in result[2] or "不銹鋼" in result[2] or "不鏽鋼" in result[2]:
            dist["eight_material"]="不鏽鋼(304)"
        if "鍍鋅" in result[2]:
            dist["eight_material"]="鍍鋅"
    temp = re.findall('厚度規號為([\d]*)[\w]*不[鏽|銹]鋼',text)
    if(len(temp) != 0):
        result = list(temp[0])
        result = [''.join(result)]
        dist["eight_material"]="不鏽鋼(304)"
        if "16" in result[0]:
            dist["eight_material"]="不鏽鋼(316)"
        if "04" in result[0]:
            dist["eight_material"]="不鏽鋼(304)"
        if "430" in result[0]:
            dist["eight_material"]="不鏽鋼(430)"
    temp = re.findall('厚度規號為([\d|\d.\d]*)[mm|mmt|mmT]*[\w]*不[鏽|銹]鋼',text)
    if(len(temp) != 0):
        result = list(temp[0])
        result = [''.join(result)]
        dist["eight_thick"]=result[0]+'t'
        dist["eight_material"]="不鏽鋼(304)"

    return dist

def BOMxlsx(filename):
    # 網址
    url = 'http://127.0.0.1:8000/media/'+str(filename)

    # 下載 Excel 檔案
    response = requests.get(url)
    # 將二進制數據轉換成 BytesIO 對象
    excel_file = BytesIO(response.content)

    # 開啟工作簿
    workbook = openpyxl.load_workbook(excel_file)

    # 選擇工作表，可以使用工作表名稱或索引（從0開始）
    # sheet = workbook['Sheet1']
    # 或者使用索引
    sheet = workbook.worksheets[0]

    # 使用 iter_cols 來獲取指定欄
    k_column = sheet.iter_cols(min_col=6, max_col=6, values_only=True)

    # 印出 K 欄的資料（避免印出 None 和 "品名規格"）
    return k_column

    # 關閉工作簿
    # workbook.close()

# def tableBOMPost(request,cNumber=None):
#     SpecBOM = Spec.objects.get(cNumber=cNumber)
#     unitinner = tableBOM.objects.filter(SpecBOM=SpecBOM).order_by("id")
#     if request.method == 'POST':
#         tablebomform = tableBOMForm(request.POST,request.FILES)
#         if tablebomform.is_valid():
#             cBOMFile = tablebomform.cleaned_data['cBOMFile']
#             chaveorder = "無"
#             tableunit = tableBOM.objects.create(SpecBOM=SpecBOM, cBOMFile=cBOMFile,chaveorder=chaveorder)
#             tableunit.save()
#             return redirect('/tableBOMPost/'+ str(SpecBOM.cNumber) +'/')
#         else:
#              message="驗證錯誤"
#     tablebomform = tableBOMForm()
#     return render(request, "tableBOMPost.html", locals())

# def tableBOMIndex(request,cNumber=None):
#     spec = Spec.objects.get(cNumber = cNumber)
#     tablebom = tableBOM.objects.filter(SpecBOM=spec).order_by('id')
#     allunittransferCount = len(tablebom)
#     return render(request, "tableBOMIndex.html",locals())

# def tableBOMDelete(request,id=None,cNumber=None):
#     spec = Spec.objects.get(cNumber = cNumber)
#     if id!=None:
#         if request.method == "POST":  #如果是以POST方式才處理
#             id=request.POST['cId'] #取得表單輸入的編號
#         try:
#             unit = tableBOM.objects.get(id=id)
#             unit.delete()
#             return redirect('/tableBOMPost/'+ str(spec.cNumber) +'/')
#         except:
#             message = "讀取錯誤!"
#     return render(request, "homeDelete.html", locals())

# def tableBOMDelete2(request,id=None,cNumber=None):
#     spec = Spec.objects.get(cNumber = cNumber)
#     if id!=None:
#         if request.method == "POST":  #如果是以POST方式才處理
#             id=request.POST['cId'] #取得表單輸入的編號
#         try:
#             unit = tableBOM.objects.get(id=id)
#             unit.delete()
#             return redirect('/tableBOMIndex/'+ str(spec.cNumber) +'/')
#         except:
#             message = "讀取錯誤!"
#     return render(request, "homeDelete.html", locals())


# def orderPost(request,id=None,cNumber=None):
#     OrderSpec = Spec.objects.get(cNumber=cNumber)
#     brands = ["高雄日月光","中壢日月光","美光","長春","旭富","天基","同步輻射","南亞電路","力成","華新科","漢欣","金像電子","喬輝","大立光","洋基","瑞昱","迅展","御勤","新竹南茂"]
#     special = "123"
#     for brand in brands:
#         if brand in OrderSpec.cClient or brand in OrderSpec.cProject:
#             print(brand)
#             try:
#                 special = Special.objects.get(cClient=brand)
#             except:
#                 special = "123"
#     unitbom = tableBOM.objects.get(id = id)
#     temp = {'machinename':' ',
#             'first_boxmaterial':' ','first_inthick':'','first_inmaterial':'','first_inother':'','first_outmaterial':'','first_outthick':'','first_boxthick':'',
#             'second_door':'','second_frame':'',
#             'third_basic':'','third_basic_install':'','third_basic_thick':'','third_basic_brand':'','third_basic_efficient':'','third_basic_frame':'','third_basic_material':'','third_medium':'','third_medium_frameform':'','third_medium_long':'','third_medium_brand':'','third_medium_efficient':'','third_medium_frame':'',
#             'third_high':'','third_high_frameform':'','third_high_long':'','third_high_brand':'','third_high_efficient':'','third_high_frame':'','third_basic_medium_material':'','third_high_material':'',
#             'four_heater':'','four_material':'','four_way':'','four_SSCR':'',
#             'fiv_humidifier':'','fiv_form':'',
#             'six_winddoor':'','six_form':'','six_frame_material':'','six_blade_material':'','six_axis':'',
#             'seven_cold':'','seven_fins':'','seven_thick':'','seven_board':'','seven_pipe_thick':'','seven_waterpipe':'','seven_size':'','seven_hot':'','seven_hotfins':'','seven_hotthick':'','seven_hotboard':'','seven_hotpipe_thick':'','seven_hotwaterpipe':'','seven_hotsize':'','seven_other':'',
#             'eight_material':'','eight_thick':'','seven_steam':'','seven_steamfins':'','seven_steamthick':'','seven_steamboard':'','seven_steampipe_thick':'','seven_steamwaterpipe':'','seven_steamsize':'','seven_brine':'','seven_brinefins':'','seven_brinethick':'','seven_brineboard':'','seven_brinepipe_thick':'','seven_brinewaterpipe':'','seven_brinesize':'',
#             'nine_brand':'科祿格','nine_way':'','nine_form':'',
#             'ten_motor_brand':'東元','ten_motor_elec':'接至端子台','ten_motor_efficient':'','ten_motor_level':'','ten_motor_open':'一般','ten_motor_cold':'自立通風','ten_motor_ask':'','ten_motor_form':'','ten_motor_Insulation':'',
#             'eleven_material':'','eleven_form':'','eleven_key':'','eleven_level':'','eleven_size':'',
#             'twelve_view':'','twelve_pressure':'','twelve_unit':'','twelve_brand':'','twelve_light':'','twelve_level':'','twelve_light_elec':'各箱段獨立電源','twelve_light_open':'','twelve_winddoor':'','twelve_net':'','televe_material':'','twelve_silicone':'勝新標準','twelve_location':'','twelve_roof':'','twelve_newmaterial':'','twelve_test':'',
#             'thirteen_way':'','thirteen_elevator':'','thirteen_box':'','thirteen_into':'','thirteen_shock':'','thirteen_shocbox':'','thirteen_apply':'','thirteen_fix':'','thirteen_budget':'','thirteen_budgetdollar':'','thirteen_displace':'無','thirteen_fixway':'無',
#             'fourteen_ass':'','fourteen_budget':'','fourteen_budgetdollar':'','fourteen_time':'','fourteen_into':'','fourteen_afterass':'','fourteen_foreigninto':'','fourteen_card':'','fourteen_insurance':'',
#             'fifteen_location':'','fifteen_level':'','fifteen_vertical':'','fifteen_box':'',
#             'sixteen_install':'','sixteen_prepare':'',
#             'seventeen_silicone':'勝新標準','seventeen_brand':'一般',
#             'eighteen_other':''}
#     fixtemp = {'third_basic_medium_material':"鍍鋅框",
# 		'eight_material':"不鏽鋼(304)",'ten_motor_elec':"接至端子台",
# 		'ten_motor_open':"一般",
# 		'ten_motor_cold':"自立通風",'twelve_silicone':"勝新標準",
# 		'twelve_light_elec':"各箱段獨立電源",'seventeen_brand':"一般"}
#     filename = unitbom.cBOMFile
#     k_column = BOMxlsx(filename)
#     dictfile = {}
#     for column in k_column:
#         for cell_value in column:
#             if cell_value is not None and cell_value != "子件編號":
#                 cell_value = re.sub(r'├', '',  cell_value)
#                 cell_value = re.sub(r'─', '',  cell_value)
#                 dictfile[cell_value] = "no"
#     a = dictfile.keys()
#     matching_products = product.objects.filter(productCode__in=a)
#     matching_productcodes = [product.productName for product in matching_products]
#     matching_productcodes = [s.replace('\t', '').replace('\n', '').replace('\r', '') for s in matching_productcodes]
#     result_dict = {key: value for item in matching_productcodes for key, value in json.loads(item).items()}
#     for key, value in result_dict.items():
#         temp[key] = value
#     temp = pressurefunc(temp,a)
#     unitbom.chaveorder = "有"
#     if request.method == 'POST':
#         orderBOMForm = OrderBOMForm(request.POST)
#         if orderBOMForm.is_valid():
#             cNumber =  unitbom.id
#             machinename=orderBOMForm.cleaned_data['machinename']
#             first_boxmaterial=orderBOMForm.cleaned_data['first_boxmaterial']
#             first_inthick=orderBOMForm.cleaned_data['first_inthick']
#             first_inmaterial=orderBOMForm.cleaned_data['first_inmaterial']
#             first_inother=orderBOMForm.cleaned_data['first_inother']
#             first_outmaterial=orderBOMForm.cleaned_data['first_outmaterial']
#             first_outthick=orderBOMForm.cleaned_data['first_outthick']
#             first_boxthick=orderBOMForm.cleaned_data['first_boxthick']
#             second_door=orderBOMForm.cleaned_data['second_door']
#             second_frame=orderBOMForm.cleaned_data['second_frame']
#             third_basic=orderBOMForm.cleaned_data['third_basic']
#             third_basic_install=orderBOMForm.cleaned_data['third_basic_install']
#             third_basic_thick=orderBOMForm.cleaned_data['third_basic_thick']
#             third_basic_brand=orderBOMForm.cleaned_data['third_basic_brand']
#             third_basic_efficient=orderBOMForm.cleaned_data['third_basic_efficient']
#             third_basic_frame=orderBOMForm.cleaned_data['third_basic_frame']
#             third_basic_material=orderBOMForm.cleaned_data['third_basic_material']
#             third_medium=orderBOMForm.cleaned_data['third_medium']
#             third_medium_frameform=orderBOMForm.cleaned_data['third_medium_frameform']
#             third_medium_long=orderBOMForm.cleaned_data['third_medium_long']
#             third_medium_brand=orderBOMForm.cleaned_data['third_medium_brand']
#             third_medium_efficient=orderBOMForm.cleaned_data['third_medium_efficient']
#             third_medium_frame=orderBOMForm.cleaned_data['third_medium_frame']
#             third_high=orderBOMForm.cleaned_data['third_high']
#             third_high_frameform=orderBOMForm.cleaned_data['third_high_frameform']
#             third_high_long=orderBOMForm.cleaned_data['third_high_long']
#             third_high_brand=orderBOMForm.cleaned_data['third_high_brand']
#             third_high_efficient=orderBOMForm.cleaned_data['third_high_efficient']
#             third_high_frame=orderBOMForm.cleaned_data['third_high_frame']
#             third_basic_medium_material=orderBOMForm.cleaned_data['third_basic_medium_material']
#             third_high_material=orderBOMForm.cleaned_data['third_high_material']
#             four_heater=orderBOMForm.cleaned_data['four_heater']
#             four_material=orderBOMForm.cleaned_data['four_material']
#             four_way=orderBOMForm.cleaned_data['four_way']
#             four_SSCR=orderBOMForm.cleaned_data['four_SSCR']
#             fiv_humidifier=orderBOMForm.cleaned_data['fiv_humidifier']
#             fiv_form=orderBOMForm.cleaned_data['fiv_form']
#             six_winddoor=orderBOMForm.cleaned_data['six_winddoor']
#             six_form=orderBOMForm.cleaned_data['six_form']
#             six_frame_material=orderBOMForm.cleaned_data['six_frame_material']
#             six_blade_material=orderBOMForm.cleaned_data['six_blade_material']
#             six_axis=orderBOMForm.cleaned_data['six_axis']
#             seven_cold=orderBOMForm.cleaned_data['seven_cold']
#             seven_fins=orderBOMForm.cleaned_data['seven_fins']
#             seven_thick=orderBOMForm.cleaned_data['seven_thick']
#             seven_board=orderBOMForm.cleaned_data['seven_board']
#             seven_pipe_thick=orderBOMForm.cleaned_data['seven_pipe_thick']
#             seven_waterpipe=orderBOMForm.cleaned_data['seven_waterpipe']
#             seven_size=orderBOMForm.cleaned_data['seven_size']
#             seven_hot=orderBOMForm.cleaned_data['seven_hot']
#             seven_hotfins=orderBOMForm.cleaned_data['seven_hotfins']
#             seven_hotthick=orderBOMForm.cleaned_data['seven_hotthick']
#             seven_hotboard=orderBOMForm.cleaned_data['seven_hotboard']
#             seven_hotpipe_thick=orderBOMForm.cleaned_data['seven_hotpipe_thick']
#             seven_hotwaterpipe=orderBOMForm.cleaned_data['seven_hotwaterpipe']
#             seven_hotsize=orderBOMForm.cleaned_data['seven_hotsize']
#             seven_steam=orderBOMForm.cleaned_data['seven_steam']
#             seven_steamfins=orderBOMForm.cleaned_data['seven_steamfins']
#             seven_steamthick=orderBOMForm.cleaned_data['seven_steamthick']
#             seven_steamboard=orderBOMForm.cleaned_data['seven_steamboard']
#             seven_steampipe_thick=orderBOMForm.cleaned_data['seven_steampipe_thick']
#             seven_steamwaterpipe=orderBOMForm.cleaned_data['seven_steamwaterpipe']
#             seven_steamsize=orderBOMForm.cleaned_data['seven_steamsize']
#             seven_brine=orderBOMForm.cleaned_data['seven_brine']
#             seven_brinefins=orderBOMForm.cleaned_data['seven_brinefins']
#             seven_brinethick=orderBOMForm.cleaned_data['seven_brinethick']
#             seven_brineboard=orderBOMForm.cleaned_data['seven_brineboard']
#             seven_brinepipe_thick=orderBOMForm.cleaned_data['seven_brinepipe_thick']
#             seven_brinewaterpipe=orderBOMForm.cleaned_data['seven_brinewaterpipe']
#             seven_brinesize=orderBOMForm.cleaned_data['seven_brinesize']
#             seven_other=orderBOMForm.cleaned_data['seven_other']
#             eight_material=orderBOMForm.cleaned_data['eight_material']
#             eight_thick=orderBOMForm.cleaned_data['eight_thick']
#             nine_brand=orderBOMForm.cleaned_data['nine_brand']
#             # nine_brand_default=orderBOMForm.cleaned_data['nine_brand_default']
#             nine_way=orderBOMForm.cleaned_data['nine_way']
#             nine_form=orderBOMForm.cleaned_data['nine_form']
#             ten_motor_brand=orderBOMForm.cleaned_data['ten_motor_brand']
#             # ten_motor_brand_default=orderBOMForm.cleaned_data['ten_motor_brand_default']
#             ten_motor_elec=orderBOMForm.cleaned_data['ten_motor_elec']
#             ten_motor_efficient=orderBOMForm.cleaned_data['ten_motor_efficient']
#             ten_motor_level=orderBOMForm.cleaned_data['ten_motor_level']
#             ten_motor_open=orderBOMForm.cleaned_data['ten_motor_open']
#             ten_motor_cold=orderBOMForm.cleaned_data['ten_motor_cold']
#             ten_motor_ask=orderBOMForm.cleaned_data['ten_motor_ask']
#             ten_motor_form=orderBOMForm.cleaned_data['ten_motor_form']
#             ten_motor_Insulation=orderBOMForm.cleaned_data['ten_motor_Insulation']
#             eleven_material=orderBOMForm.cleaned_data['eleven_material']
#             eleven_form=orderBOMForm.cleaned_data['eleven_form']
#             eleven_key=orderBOMForm.cleaned_data['eleven_key']
#             eleven_level=orderBOMForm.cleaned_data['eleven_level']
#             eleven_size=orderBOMForm.cleaned_data['eleven_size']
#             twelve_view=orderBOMForm.cleaned_data['twelve_view']
#             twelve_pressure=orderBOMForm.cleaned_data['twelve_pressure']
#             twelve_unit=orderBOMForm.cleaned_data['twelve_unit']
#             twelve_brand=orderBOMForm.cleaned_data['twelve_brand']
#             twelve_light=orderBOMForm.cleaned_data['twelve_light']
#             twelve_level=orderBOMForm.cleaned_data['twelve_level']
#             twelve_light_elec=orderBOMForm.cleaned_data['twelve_light_elec']
#             twelve_light_open=orderBOMForm.cleaned_data['twelve_light_open']
#             twelve_winddoor=orderBOMForm.cleaned_data['twelve_winddoor']
#             twelve_net=orderBOMForm.cleaned_data['twelve_net']
#             twelve_material=orderBOMForm.cleaned_data['twelve_material']
#             twelve_silicone=orderBOMForm.cleaned_data['twelve_silicone']
#             # twelve_silicone_default=orderBOMForm.cleaned_data['twelve_silicone_default']
#             twelve_location=orderBOMForm.cleaned_data['twelve_location']
#             twelve_roof=orderBOMForm.cleaned_data['twelve_roof']
#             twelve_newmaterial=orderBOMForm.cleaned_data['twelve_newmaterial']
#             twelve_test=orderBOMForm.cleaned_data['twelve_test']
#             thirteen_way=orderBOMForm.cleaned_data['thirteen_way']
#             thirteen_elevator=orderBOMForm.cleaned_data['thirteen_elevator']
#             thirteen_box=orderBOMForm.cleaned_data['thirteen_box']
#             thirteen_into=orderBOMForm.cleaned_data['thirteen_into']
#             thirteen_shock=orderBOMForm.cleaned_data['thirteen_shock']
#             thirteen_shocbox=orderBOMForm.cleaned_data['thirteen_shocbox']
#             thirteen_apply=orderBOMForm.cleaned_data['thirteen_apply']
#             thirteen_fix=orderBOMForm.cleaned_data['thirteen_fix']
#             thirteen_budget=orderBOMForm.cleaned_data['thirteen_budget']
#             thirteen_budgetdollar=orderBOMForm.cleaned_data['thirteen_budgetdollar']
#             thirteen_displace=orderBOMForm.cleaned_data['thirteen_displace']
#             # thirteen_displace_default=orderBOMForm.cleaned_data['thirteen_displace_default']
#             thirteen_fixway=orderBOMForm.cleaned_data['thirteen_fixway']
#             # thirteen_fixway_default=orderBOMForm.cleaned_data['thirteen_fixway_default']
#             fourteen_ass=orderBOMForm.cleaned_data['fourteen_ass']
#             fourteen_budget=orderBOMForm.cleaned_data['fourteen_budget']
#             fourteen_budgetdollar=orderBOMForm.cleaned_data['fourteen_budgetdollar']
#             fourteen_time=orderBOMForm.cleaned_data['fourteen_time']
#             fourteen_into=orderBOMForm.cleaned_data['fourteen_into']
#             fourteen_intotime=orderBOMForm.cleaned_data['fourteen_intotime']
#             fourteen_afterass=orderBOMForm.cleaned_data['fourteen_afterass']
#             fourteen_afterasstime=orderBOMForm.cleaned_data['fourteen_afterasstime']
#             fourteen_foreigninto=orderBOMForm.cleaned_data['fourteen_foreigninto']
#             fourteen_foreignintotime=orderBOMForm.cleaned_data['fourteen_foreignintotime']
#             fourteen_card=orderBOMForm.cleaned_data['fourteen_card']
#             fourteen_insurance=orderBOMForm.cleaned_data['fourteen_insurance']
#             fifteen_location=orderBOMForm.cleaned_data['fifteen_location']
#             fifteen_level=orderBOMForm.cleaned_data['fifteen_level']
#             fifteen_vertical=orderBOMForm.cleaned_data['fifteen_vertical']
#             fifteen_box=orderBOMForm.cleaned_data['fifteen_box']
#             sixteen_install=orderBOMForm.cleaned_data['sixteen_install']
#             sixteen_prepare=orderBOMForm.cleaned_data['sixteen_prepare']
#             seventeen_silicone=orderBOMForm.cleaned_data['seventeen_silicone']
#             seventeen_brand=orderBOMForm.cleaned_data['seventeen_brand']
#             eighteen_other=orderBOMForm.cleaned_data['eighteen_other']


#             OrderBOMunit = OrderBOM.objects.create(home=unitbom, cNumber=cNumber, 
#                         machinename=machinename,
#                         first_boxmaterial=first_boxmaterial,
#                         first_inthick=first_inthick,
#                         first_inmaterial=first_inmaterial,
#                         first_inother=first_inother,
#                         first_outmaterial=first_outmaterial,
#                         first_outthick=first_outthick,
#                         first_boxthick=first_boxthick,
#                         second_door=second_door,
#                         second_frame=second_frame,
#                         third_basic=third_basic,
#                         third_basic_install=third_basic_install,
#                         third_basic_thick=third_basic_thick,
#                         third_basic_brand=third_basic_brand,
#                         third_basic_efficient=third_basic_efficient,
#                         third_basic_frame=third_basic_frame,
#                         third_basic_material=third_basic_material,
#                         third_medium=third_medium,
#                         third_medium_frameform=third_medium_frameform,
#                         third_medium_long=third_medium_long,
#                         third_medium_brand=third_medium_brand,
#                         third_medium_efficient=third_medium_efficient,
#                         third_medium_frame=third_medium_frame,
#                         third_high=third_high,
#                         third_high_frameform=third_high_frameform,
#                         third_high_long=third_high_long,
#                         third_high_brand=third_high_brand,
#                         third_high_efficient=third_high_efficient,
#                         third_high_frame=third_high_frame,
#                         third_high_material=third_high_material,
#                         third_basic_medium_material=third_basic_medium_material,
#                         four_heater=four_heater,
#                         four_material=four_material,
#                         four_way=four_way,
#                         four_SSCR=four_SSCR,
#                         fiv_humidifier=fiv_humidifier,
#                         fiv_form=fiv_form,
#                         six_winddoor=six_winddoor,
#                         six_form=six_form,
#                         six_frame_material=six_frame_material,
#                         six_blade_material=six_blade_material,
#                         six_axis=six_axis,
#                         seven_cold=seven_cold,
#                         seven_fins=seven_fins,
#                         seven_thick=seven_thick,
#                         seven_board=seven_board,
#                         seven_pipe_thick=seven_pipe_thick,
#                         seven_waterpipe=seven_waterpipe,
#                         seven_size=seven_size,
#                         seven_other=seven_other,
#                         seven_hot=seven_hot,
#                         seven_hotfins=seven_hotfins,
#                         seven_hotthick=seven_hotthick,
#                         seven_hotboard=seven_hotboard,
#                         seven_hotpipe_thick=seven_hotpipe_thick,
#                         seven_hotwaterpipe=seven_hotwaterpipe,
#                         seven_hotsize=seven_hotsize,
#                         seven_steam=seven_steam,
#                         seven_steamfins=seven_steamfins,
#                         seven_steamthick=seven_steamthick,
#                         seven_steamboard=seven_steamboard,
#                         seven_steampipe_thick=seven_steampipe_thick,
#                         seven_steamwaterpipe=seven_steamwaterpipe,
#                         seven_steamsize=seven_steamsize,
#                         seven_brine=seven_brine,
#                         seven_brinefins=seven_brinefins,
#                         seven_brinethick=seven_brinethick,
#                         seven_brineboard=seven_brineboard,
#                         seven_brinepipe_thick=seven_brinepipe_thick,
#                         seven_brinewaterpipe=seven_brinewaterpipe,
#                         seven_brinesize=seven_brinesize,
#                         eight_material=eight_material,
#                         eight_thick=eight_thick,
#                         nine_brand=nine_brand,
#                         nine_way=nine_way,
#                         nine_form=nine_form,
#                         ten_motor_brand=ten_motor_brand,
#                         ten_motor_elec=ten_motor_elec,
#                         ten_motor_efficient=ten_motor_efficient,
#                         ten_motor_level=ten_motor_level,
#                         ten_motor_open=ten_motor_open,
#                         ten_motor_cold=ten_motor_cold,
#                         ten_motor_ask=ten_motor_ask,
#                         ten_motor_form=ten_motor_form,
#                         ten_motor_Insulation=ten_motor_Insulation,
#                         eleven_material=eleven_material,
#                         eleven_form=eleven_form,
#                         eleven_key=eleven_key,
#                         eleven_level=eleven_level,
#                         eleven_size=eleven_size,
#                         twelve_view=twelve_view,
#                         twelve_pressure=twelve_pressure,
#                         twelve_unit=twelve_unit,
#                         twelve_brand=twelve_brand,
#                         twelve_light=twelve_light,
#                         twelve_level=twelve_level,
#                         twelve_light_elec=twelve_light_elec,
#                         twelve_light_open=twelve_light_open,
#                         twelve_winddoor=twelve_winddoor,
#                         twelve_net=twelve_net,
#                         twelve_material=twelve_material,
#                         twelve_silicone=twelve_silicone,
#                         twelve_location=twelve_location,
#                         twelve_roof=twelve_roof,
#                         twelve_newmaterial=twelve_newmaterial,
#                         twelve_test=twelve_test,
#                         thirteen_way=thirteen_way,
#                         thirteen_elevator=thirteen_elevator,
#                         thirteen_box=thirteen_box,
#                         thirteen_into=thirteen_into,
#                         thirteen_shock=thirteen_shock,
#                         thirteen_shocbox=thirteen_shocbox,
#                         thirteen_apply=thirteen_apply,
#                         thirteen_fix=thirteen_fix,
#                         thirteen_budget=thirteen_budget,
#                         thirteen_budgetdollar=thirteen_budgetdollar,
#                         thirteen_displace=thirteen_displace,
#                         thirteen_fixway=thirteen_fixway,
#                         fourteen_ass=fourteen_ass,
#                         fourteen_budget=fourteen_budget,
#                         fourteen_budgetdollar=fourteen_budgetdollar,
#                         fourteen_time=fourteen_time,
#                         fourteen_into=fourteen_into,
#                         fourteen_intotime=fourteen_intotime,
#                         fourteen_afterass=fourteen_afterass,
#                         fourteen_afterasstime=fourteen_afterasstime,
#                         fourteen_foreigninto=fourteen_foreigninto,
#                         fourteen_foreignintotime=fourteen_foreignintotime,
#                         fourteen_card=fourteen_card,
#                         fourteen_insurance=fourteen_insurance,
#                         fifteen_location=fifteen_location,
#                         fifteen_level=fifteen_level,
#                         fifteen_vertical=fifteen_vertical,
#                         fifteen_box=fifteen_box,
#                         sixteen_install=sixteen_install,
#                         sixteen_prepare=sixteen_prepare,
#                         seventeen_silicone=seventeen_silicone,
#                         seventeen_brand=seventeen_brand,
#                         eighteen_other=eighteen_other)
#             OrderBOMunit.save()
#             unitbom.save()
#             return redirect('/orderView/'+str(cNumber)+'/'+str(OrderSpec.cNumber)+'/')
#         else:
#              message="驗證錯誤"
#     else:
#          message=''
#     orderBOMForm = OrderBOMForm(temp)
#     return render(request, "orderPost.html", locals())




# def orderView(request,id=None,cNumber=None):
#     if 'user_id' in request.session:  # 假设你的用户 ID 存在于 session 中
#         user_id = request.session['user_id']
#         user = User.objects.get(pk=user_id)  # 获取当前登录用户信息
#         firstname = user.name  # 传递用户名到模板中
#         position = user.Position
#         title = user.title
#     allPerson = Home.objects.filter(cReceive = firstname)
#     allpersonCount = len(allPerson)
#     OrderSpec = Spec.objects.get(cNumber=cNumber)
#     home = tableBOM.objects.get(id=id)
#     # returned = Returned.objects.filter(returnTo=home).order_by("id")
#     order = OrderBOM.objects.get(cNumber=id)
#     return render(request, "orderView.html",locals()) 

# def orderEdit(request,mode=None,id=None,cNumber=None):
#     OrderSpec = Spec.objects.get(cNumber=cNumber)
#     orderBOMform = OrderBOMForm(request.POST)
#     if mode == "load":  # 由 index.html 按 編輯二 鈕
#         orderBOM = OrderBOM.objects.get(cNumber = id)  #取得要修改的資料記
#         return render(request, "orderEdit.html", locals())
#     elif mode == "save": # 由 edit2.html 按 submit
#         orderBOM = OrderBOM.objects.get(cNumber = id)  #取得要修改的資料記錄
#         orderBOM.machinename=request.POST['machinename']
#         orderBOM.first_boxmaterial=request.POST['first_boxmaterial']
#         orderBOM.first_inthick=request.POST['first_inthick']
#         orderBOM.first_inmaterial=request.POST['first_inmaterial']
#         orderBOM.first_inother=request.POST['first_inother']
#         orderBOM.first_outmaterial=request.POST['first_outmaterial']
#         orderBOM.first_outthick=request.POST['first_outthick']
#         orderBOM.first_boxthick=request.POST['first_boxthick']
#         orderBOM.second_door=request.POST['second_door']
#         orderBOM.second_frame=request.POST['second_frame']
#         orderBOM.third_basic=request.POST['third_basic']
#         orderBOM.third_basic_install=request.POST['third_basic_install']
#         orderBOM.third_basic_thick=request.POST['third_basic_thick']
#         orderBOM.third_basic_brand=request.POST['third_basic_brand']
#         orderBOM.third_basic_efficient=request.POST['third_basic_efficient']
#         orderBOM.third_basic_frame=request.POST['third_basic_frame']
#         orderBOM.third_basic_material=request.POST['third_basic_material']
#         orderBOM.third_medium=request.POST['third_medium']
#         orderBOM.third_medium_frameform=request.POST['third_medium_frameform']
#         orderBOM.third_medium_long=request.POST['third_medium_long']
#         orderBOM.third_medium_brand=request.POST['third_medium_brand']
#         orderBOM.third_medium_efficient=request.POST['third_medium_efficient']
#         orderBOM.third_medium_frame=request.POST['third_medium_frame']
#         orderBOM.third_high=request.POST['third_high']
#         orderBOM.third_high_frameform=request.POST['third_high_frameform']
#         orderBOM.third_high_long=request.POST['third_high_long']
#         orderBOM.third_high_brand=request.POST['third_high_brand']
#         orderBOM.third_high_efficient=request.POST['third_high_efficient']
#         orderBOM.third_high_frame=request.POST['third_high_frame']
#         orderBOM.third_basic_medium_material=request.POST['third_basic_medium_material']
#         orderBOM.third_high_material=request.POST['third_high_material']
#         orderBOM.four_heater=request.POST['four_heater']
#         orderBOM.four_material=request.POST['four_material']
#         orderBOM.four_way=request.POST['four_way']
#         orderBOM.four_SSCR=request.POST['four_SSCR']
#         orderBOM.fiv_humidifier=request.POST['fiv_humidifier']
#         orderBOM.form=request.POST['fiv_form']
#         orderBOM.six_winddoor=request.POST['six_winddoor']
#         orderBOM.six_form=request.POST['six_form']
#         orderBOM.six_frame_material=request.POST['six_frame_material']
#         orderBOM.six_blade_material=request.POST['six_blade_material']
#         orderBOM.six_axis=request.POST['six_axis']
#         orderBOM.seven_cold=request.POST['seven_cold']
#         orderBOM.seven_fins=request.POST['seven_fins']
#         orderBOM.seven_thick=request.POST['seven_thick']
#         orderBOM.seven_board=request.POST['seven_board']
#         orderBOM.seven_pipe_thick=request.POST['seven_pipe_thick']
#         orderBOM.seven_waterpipe=request.POST['seven_waterpipe']
#         orderBOM.seven_size=request.POST['seven_size']
#         orderBOM.seven_hot=request.POST['seven_hot']
#         orderBOM.seven_hotfins=request.POST['seven_hotfins']
#         orderBOM.seven_hotthick=request.POST['seven_hotthick']
#         orderBOM.seven_hotboard=request.POST['seven_hotboard']
#         orderBOM.seven_hotpipe_thick=request.POST['seven_hotpipe_thick']
#         orderBOM.seven_hotwaterpipe=request.POST['seven_hotwaterpipe']
#         orderBOM.seven_hotsize=request.POST['seven_hotsize']
#         orderBOM.seven_steam=request.POST['seven_steam']
#         orderBOM.seven_steamfins=request.POST['seven_steamfins']
#         orderBOM.seven_steamthick=request.POST['seven_steamthick']
#         orderBOM.seven_steamboard=request.POST['seven_steamboard']
#         orderBOM.seven_steampipe_thick=request.POST['seven_steampipe_thick']
#         orderBOM.seven_steamwaterpipe=request.POST['seven_steamwaterpipe']
#         orderBOM.seven_steamsize=request.POST['seven_steamsize']
#         orderBOM.seven_brine=request.POST['seven_brine']
#         orderBOM.seven_brinefins=request.POST['seven_brinefins']
#         orderBOM.seven_brinethick=request.POST['seven_brinethick']
#         orderBOM.seven_brineboard=request.POST['seven_brineboard']
#         orderBOM.seven_brinepipe_thick=request.POST['seven_brinepipe_thick']
#         orderBOM.seven_brinewaterpipe=request.POST['seven_brinewaterpipe']
#         orderBOM.seven_brinesize=request.POST['seven_brinesize']
#         orderBOM.seven_other=request.POST['seven_other']
#         orderBOM.eight_material=request.POST['eight_material']
#         orderBOM.eight_thick=request.POST['eight_thick']
#         orderBOM.nine_brand=request.POST['nine_brand']
#         orderBOM.nine_way=request.POST['nine_way']
#         orderBOM.nine_form=request.POST['nine_form']
#         orderBOM.ten_motor_brand=request.POST['ten_motor_brand']
#         orderBOM.ten_motor_elec=request.POST['ten_motor_elec']
#         orderBOM.ten_motor_efficient=request.POST['ten_motor_efficient']
#         orderBOM.ten_motor_level=request.POST['ten_motor_level']
#         orderBOM.ten_motor_open=request.POST['ten_motor_open']
#         orderBOM.ten_motor_cold=request.POST['ten_motor_cold']
#         orderBOM.ten_motor_ask=request.POST['ten_motor_ask']
#         orderBOM.ten_motor_form=request.POST['ten_motor_form']
#         orderBOM.ten_motor_Insulation=request.POST['ten_motor_form']
#         orderBOM.eleven_material=request.POST['eleven_material']
#         orderBOM.eleven_form=request.POST['eleven_form']
#         orderBOM.eleven_key=request.POST['eleven_key']
#         orderBOM.eleven_level=request.POST['eleven_level']
#         orderBOM.eleven_size=request.POST['eleven_size']
#         orderBOM.twelve_view=request.POST['twelve_view']
#         orderBOM.twelve_pressure=request.POST['twelve_pressure']
#         orderBOM.twelve_unit=request.POST['twelve_unit']
#         orderBOM.twelve_brand=request.POST['twelve_brand']
#         orderBOM.twelve_light=request.POST['twelve_light']
#         orderBOM.twelve_level=request.POST['twelve_level']
#         orderBOM.twelve_light_elec=request.POST['twelve_light_elec']
#         orderBOM.twelve_light_open=request.POST['twelve_light_open']
#         orderBOM.twelve_winddoor=request.POST['twelve_winddoor']
#         orderBOM.twelve_net=request.POST['twelve_net']
#         orderBOM.twelve_material=request.POST['twelve_material']
#         orderBOM.twelve_silicone=request.POST['twelve_silicone']
#         orderBOM.twelve_roof=request.POST['twelve_roof']
#         orderBOM.twelve_newmaterial=request.POST['twelve_newmaterial']
#         orderBOM.twelve_location=request.POST['twelve_location']
#         orderBOM.twelve_test=request.POST['twelve_test']
#         orderBOM.thirteen_way=request.POST['thirteen_way']
#         orderBOM.thirteen_elevator=request.POST['thirteen_elevator']
#         orderBOM.thirteen_box=request.POST['thirteen_box']
#         orderBOM.thirteen_into=request.POST['thirteen_into']
#         orderBOM.thirteen_shock=request.POST['thirteen_shock']
#         orderBOM.thirteen_shocbox=request.POST['thirteen_shocbox']
#         orderBOM.thirteen_apply=request.POST['thirteen_apply']
#         orderBOM.thirteen_fix=request.POST['thirteen_fix']
#         orderBOM.thirteen_budget=request.POST['thirteen_budget']
#         orderBOM.thirteen_displace=request.POST['thirteen_displace']
#         orderBOM.thirteen_fixway=request.POST['thirteen_fixway']
#         orderBOM.fourteen_ass=request.POST['fourteen_ass']
#         orderBOM.fourteen_budget=request.POST['fourteen_budget']
#         orderBOM.fourteen_time=request.POST['fourteen_time']
#         orderBOM.fourteen_into=request.POST['fourteen_into']
#         orderBOM.fourteen_intotime=request.POST['fourteen_intotime']
#         orderBOM.fourteen_afterass=request.POST['fourteen_afterass']
#         orderBOM.fourteen_afterasstime=request.POST['fourteen_afterasstime']
#         orderBOM.fourteen_foreigninto=request.POST['fourteen_foreigninto']
#         orderBOM.fourteen_foreignintotime=request.POST['fourteen_foreignintotime']
#         orderBOM.fourteen_card=request.POST['fourteen_card']
#         orderBOM.fourteen_insurance=request.POST['fourteen_insurance']
#         orderBOM.fifteen_location=request.POST['fifteen_location']
#         orderBOM.fifteen_level=request.POST['fifteen_level']
#         orderBOM.fifteen_vertical=request.POST['fifteen_vertical']
#         orderBOM.fifteen_box=request.POST['fifteen_box']
#         orderBOM.sixteen_install=request.POST['sixteen_install']
#         orderBOM.sixteen_prepare=request.POST['sixteen_prepare']
#         orderBOM.seventeen_silicone=request.POST['seventeen_silicone']
#         orderBOM.seventeen_brand=request.POST['seventeen_brand']
#         orderBOM.eighteen_other=request.POST['eighteen_other']
#         orderBOM.save()  #寫入資料庫
#           #寫入資料庫
#         return redirect('/orderView/'+str(id)+"/"+str(OrderSpec.cNumber)+'/')
    


# def pressurefunc(temp,array):
#     for i in array:
#         if(i.startswith("S13")):
#             temp['twelve_pressure'] = "指針式X1"
#             temp['twelve_unit'] = "mmAq"
#             continue
#         elif(i.startswith("S12")):
#             temp['twelve_light'] = "有"
#             continue
#         elif(i.startswith("S9")):
#             temp['twelve_view'] = "有"
#             continue
#     return temp



    